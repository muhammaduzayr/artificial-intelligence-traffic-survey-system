"""
Exports survey results directly into a COPY of the official counting
template (Counting_Custom_Rev1.xlsx, path in config.TMC_TEMPLATE_PATH),
instead of generating a workbook from scratch.

This matters because the template isn't just cell values — each of its
five sheets (Direction, North, East, South, West) carries its own
embedded vector drawing (the little turn-arrow icons next to each
movement). A plain "load template with openpyxl, edit cells, save"
approach SILENTLY DROPS all of those drawings on save — tested and
confirmed. So instead, this module edits the underlying worksheet XML
directly inside a raw copy of the template .xlsx (which is just a zip
archive), touching only the specific <c> cell elements that need new
values. Every other part of the archive — drawings, styles, merged
cells, the Direction sheet, printer settings — is copied through
byte-for-byte unchanged.

Sheet layout this assumes (matches Counting_Custom_Rev1.xlsx exactly):
  - One sheet per arm: "North", "East", "South", "West"
  - B2/B3/B4/B5 = Enumerator/Junction/Junction Name/Date LABELS (merged
    B_:F_); the VALUES go in the merged G_:Z_ cell alongside each label
    (the top-left cell of that merge, G, is where OOXML actually stores
    the value).
  - Row 7: movement code headers (e.g. N1/N2/N3/N4 on the North sheet),
    one per movement block, merged across that block's 5 columns.
  - Row 9: C / L / LL / B / M sub-column headers within each block.
  - Movement block starting columns: 1->C(3), 2->I(9), 3->O(15), 4->U(21)
  - Time column B, rows 10-74, ALREADY pre-filled 06:00-22:00 in 15-min
    steps by the template itself — we look up the matching row for each
    interval rather than writing our own time column, so this keeps
    working even if a template variant uses a slightly different range.
"""

import os
import re
import shutil
import tempfile
import zipfile
from xml.sax.saxutils import escape

import openpyxl
from openpyxl.utils import get_column_letter

from . import config


# Map our internal category names to the template's sub-column codes.
# Refined in config.py: COCO class 7 splits into Light Truck / Truck
# by bbox area threshold.
CATEGORY_TO_COLUMN = {
    "Car": "C",
    "Light Truck": "L",
    "Truck": "LL",
    "Bus": "B",
    "Motorcycle": "M",
}

def check_raw_df_for_double_counts(raw_df, cooldown_sec):
    """
    Export-time self-check: inspect raw_events for any track_id that
    appears more than once within cooldown_sec.  If found, log a loud
    warning with both event details — this catches both legitimate
    multi-zone crossings (same vehicle counted in two zones) and any
    actual ID-reuse bugs, forcing manual review instead of silently
    double-counting.
    """
    if raw_df.empty or "track_id" not in raw_df.columns:
        return

    df = raw_df.sort_values(["track_id", "interval_start"]).copy()
    df["prev_lane"] = df.groupby("track_id")["lane"].shift(1)
    df["prev_time"] = df.groupby("track_id")["interval_start"].shift(1)
    df["time_gap_sec"] = (df["interval_start"] - df["prev_time"]).dt.total_seconds()

    close_pairs = df[
        (df["time_gap_sec"].notna())
        & (df["time_gap_sec"] <= cooldown_sec)
    ]
    if not close_pairs.empty:
        print("=" * 72)
        print("  DOUBLE-COUNT WARNING (track_id appears twice within cooldown)")
        print("=" * 72)
        for _, row in close_pairs.iterrows():
            prev_lane = row["prev_lane"]
            cur_lane = row["lane"]
            same_lane = prev_lane == cur_lane
            note = "SAME lane - true double-count risk" if same_lane else "DIFFERENT lanes - multi-zone crossing"
            print(
                f"  track_id={int(row['track_id'])}  "
                f"gap={row['time_gap_sec']:.0f}s  "
                f"{prev_lane} -> {cur_lane}  "
                f"at {row['interval_start'].strftime('%H:%M')}  "
                f"[{note}]"
            )
        print("=" * 72)


ARM_SHEET_NAMES = {"N": "North", "E": "East", "S": "South", "W": "West"}

# Column index (1-based) of the first ("C") sub-column for each movement block.
MOVEMENT_BLOCK_START_COL = {1: 3, 2: 9, 3: 15, 4: 21}
SUBCOL_OFFSET = {"C": 0, "L": 1, "LL": 2, "B": 3, "M": 4}

# Row of each header LABEL (Enumerator/Junction/Junction Name/Date); the
# matching VALUE cell sits in the same row, at VALUE_COL (the anchor of
# the merged G_:Z_ range next to each label).
HEADER_LABEL_ROWS = {"enumerator": 2, "junction": 3, "junction_name": 4, "date": 5}
VALUE_COL = 7  # column G

FIRST_TIME_ROW = 10
LAST_TIME_ROW = 74  # matches the template's fixed 06:00-22:00 range


def parse_zone_label(label):
    """'N1' -> ('N', 1). Returns (None, None) if the label doesn't match
    the <Direction><Movement 1-4> pattern this template requires."""
    if not isinstance(label, str):
        return None, None
    m = re.match(r"^\s*([NSEW])\s*(\d)\s*$", label.strip().upper())
    if not m:
        return None, None
    movement = int(m.group(2))
    if movement not in MOVEMENT_BLOCK_START_COL:
        return None, None
    return m.group(1), movement


def _cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


def _time_to_row(lookup_wb, sheet_name):
    """Reads the template's existing time column (B10:B74) and builds a
    {"HH:MM" -> row number} map, so counts get matched to the row the
    template ALREADY has for that time, rather than assuming a fixed
    pattern in code."""
    ws = lookup_wb[sheet_name]
    mapping = {}
    for row in range(FIRST_TIME_ROW, LAST_TIME_ROW + 1):
        v = ws.cell(row=row, column=2).value
        if v is None:
            continue
        key = v.strftime("%H:%M") if hasattr(v, "strftime") else str(v)
        mapping[key] = row
    return mapping


def _build_cell_xml(ref, style_attr, kind, value):
    """kind: "str" (written as an inline string, no sharedStrings.xml
    changes needed) or "num" (plain numeric)."""
    style_part = f' s="{style_attr}"' if style_attr else ""
    if kind == "str":
        safe = escape(str(value))
        return f'<c r="{ref}"{style_part} t="inlineStr"><is><t xml:space="preserve">{safe}</t></is></c>'
    return f'<c r="{ref}"{style_part}><v>{value}</v></c>'


def _col_letters_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _set_cells_in_sheet_xml(xml_bytes, cell_values):
    """
    cell_values: {cell_ref: (kind, value)}, e.g. {"G2": ("str", "Aiman")}

    Directly rewrites specific <c> elements inside one worksheet XML
    part's <sheetData>, preserving that cell's existing style (s="N")
    attribute if it had one, and leaving everything else in the sheet —
    and every OTHER file in the archive, including this sheet's own
    <drawing> reference — completely untouched.
    """
    text = xml_bytes.decode("utf-8")

    for ref, (kind, value) in cell_values.items():
        row_num = int(re.match(r"[A-Z]+(\d+)", ref).group(1))

        # Try to find and replace an existing <c r="REF" .../> or
        # <c r="REF">...</c> — the template's cells are pre-styled
        # (borders/merges), so this is the path almost every write hits.
        pattern = re.compile(rf'<c r="{ref}"([^>]*?)(?:/>|>(.*?)</c>)', re.DOTALL)
        match = pattern.search(text)
        if match:
            attrs = match.group(1) or ""
            style_match = re.search(r's="(\d+)"', attrs)
            style_attr = style_match.group(1) if style_match else None
            new_cell_xml = _build_cell_xml(ref, style_attr, kind, value)
            text = text[:match.start()] + new_cell_xml + text[match.end():]
            continue

        # Fallback: cell doesn't exist in the XML yet (rare — only if the
        # template left a cell fully absent rather than an empty styled
        # placeholder). Insert it into its row in column order, or create
        # the row itself if that's missing too.
        new_cell_xml = _build_cell_xml(ref, None, kind, value)
        row_pattern = re.compile(rf'(<row r="{row_num}"[^>]*>)(.*?)(</row>)', re.DOTALL)
        row_match = row_pattern.search(text)
        if row_match:
            row_open, row_body, row_close = row_match.groups()
            insert_col = re.match(r"[A-Z]+", ref).group(0)
            insert_pos = len(row_body)
            for m in re.finditer(r'<c r="([A-Z]+)\d+"[^>]*?(?:/>|>.*?</c>)', row_body, re.DOTALL):
                if _col_letters_to_num(m.group(1)) > _col_letters_to_num(insert_col):
                    insert_pos = m.start()
                    break
            new_row_body = row_body[:insert_pos] + new_cell_xml + row_body[insert_pos:]
            text = text[:row_match.start()] + row_open + new_row_body + row_close + text[row_match.end():]
        else:
            new_row = f'<row r="{row_num}">{new_cell_xml}</row>'
            text = text.replace("</sheetData>", new_row + "</sheetData>", 1)

    return text.encode("utf-8")


def _map_sheet_names_to_files(zf):
    """Reads xl/workbook.xml + xl/_rels/workbook.xml.rels to figure out
    which xl/worksheets/sheetN.xml file corresponds to each sheet NAME —
    sheet order/ids in the workbook aren't guaranteed to match the
    sheetN.xml file numbering."""
    workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
    rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    name_to_rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', workbook_xml))
    rid_to_target = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))

    result = {}
    for name, rid in name_to_rid.items():
        target = rid_to_target[rid]  # e.g. "worksheets/sheet2.xml"
        result[name] = "xl/" + target.lstrip("/")
    return result


def _patch_zip_in_place(xlsx_path, edits_by_sheet_file):
    """Rewrites only the specified worksheet XML parts inside the given
    .xlsx, copying every other part of the archive through unmodified —
    this is what keeps the template's drawings/styles/other sheets
    byte-for-byte identical to the original."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in edits_by_sheet_file:
                data = _set_cells_in_sheet_xml(data, edits_by_sheet_file[item.filename])
            zout.writestr(item, data)

    shutil.move(tmp_path, xlsx_path)


def export_tmc_report(raw_df, survey_start_time, output_path, interval_minutes=15,
                       enumerator="", junction="", junction_name="", date_str=None):
    """
    raw_df: aggregator.raw_dataframe() — needs columns interval_start,
        lane, category. `lane` values should look like "N1", "E2", etc.
    survey_start_time: used as a fallback if raw_df is empty, and to
        derive the default date string.

    Writes a COPY of config.TMC_TEMPLATE_PATH to output_path with the
    header fields and count cells filled in. Every other part of the
    template — styling, the Direction sheet, per-sheet arrow diagrams,
    merged cells — is preserved exactly as-is.

    Returns (output_path, unplaced_labels) — unplaced_labels lists any
    zone labels found in the data that didn't match the N/E/S/W + 1-4
    pattern and so couldn't be placed anywhere in this template.
    """
    if not os.path.exists(config.TMC_TEMPLATE_PATH):
        raise FileNotFoundError(
            f"TMC template not found at {config.TMC_TEMPLATE_PATH}. Copy "
            f"Counting_Custom_Rev1.xlsx there, or update config.TMC_TEMPLATE_PATH."
        )

    if date_str is None:
        date_str = survey_start_time.strftime("%Y-%m-%d")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    shutil.copyfile(config.TMC_TEMPLATE_PATH, output_path)

    # Read-only pass, purely to look up which row each 15-min interval
    # belongs to in the template's pre-filled time column. This workbook
    # object is never saved — only the raw zip copy made above (output_path)
    # gets written to disk, via the XML patching below.
    lookup_wb = openpyxl.load_workbook(config.TMC_TEMPLATE_PATH, data_only=True)

    unplaced_labels = set()
    if not raw_df.empty:
        parsed = raw_df["lane"].map(parse_zone_label)
        for label, (arm, movement) in zip(raw_df["lane"], parsed):
            if arm is None:
                unplaced_labels.add(label)

    with zipfile.ZipFile(config.TMC_TEMPLATE_PATH) as zf:
        sheet_name_to_file = _map_sheet_names_to_files(zf)

    edits_by_sheet_file = {}  # "xl/worksheets/sheetN.xml" -> {ref: (kind, value)}

    for arm_letter, sheet_name in ARM_SHEET_NAMES.items():
        sheet_file = sheet_name_to_file[sheet_name]
        cell_edits = {
            _cell_ref(HEADER_LABEL_ROWS["enumerator"], VALUE_COL): ("str", enumerator),
            _cell_ref(HEADER_LABEL_ROWS["junction"], VALUE_COL): ("str", junction),
            _cell_ref(HEADER_LABEL_ROWS["junction_name"], VALUE_COL): ("str", junction_name),
            _cell_ref(HEADER_LABEL_ROWS["date"], VALUE_COL): ("str", date_str),
        }

        if not raw_df.empty:
            time_to_row = _time_to_row(lookup_wb, sheet_name)
            arm_df = raw_df.copy()
            parsed = arm_df["lane"].map(parse_zone_label)
            arm_df["arm"] = [p[0] for p in parsed]
            arm_df["movement"] = [p[1] for p in parsed]
            arm_df = arm_df[arm_df["arm"] == arm_letter]

            if not arm_df.empty:
                grouped = arm_df.groupby(["interval_start", "movement", "category"]).size()
                total_placed = 0
                total_skipped_category = 0
                total_skipped_time = 0
                for (interval_start, movement, category), count in grouped.items():
                    subcol = CATEGORY_TO_COLUMN.get(category)
                    if subcol is None:
                        total_skipped_category += 1
                        if total_skipped_category == 1:
                            print(f"[tmc-export] WARNING: no subcolumn mapping for "
                                  f"category={category!r} (sheet={sheet_name})")
                        continue
                    row_idx = time_to_row.get(interval_start.strftime("%H:%M"))
                    if row_idx is None:
                        total_skipped_time += 1
                        if total_skipped_time == 1:
                            print(f"[tmc-export] WARNING: time {interval_start.strftime('%H:%M')} "
                                  f"not found in {sheet_name} time column (outside 06:00-22:00)")
                        continue
                    col_idx = MOVEMENT_BLOCK_START_COL[movement] + SUBCOL_OFFSET[subcol]
                    ref = _cell_ref(row_idx, col_idx)
                    cell_edits[ref] = ("num", int(count))
                    total_placed += 1
                if not arm_df.empty:
                    placed_info = f" placed={total_placed}"
                    if total_skipped_category:
                        placed_info += f" skip-cat={total_skipped_category}"
                    if total_skipped_time:
                        placed_info += f" skip-time={total_skipped_time}"
                    print(f"[tmc-export] {sheet_name}: {len(arm_df)} rows{placed_info}")

        edits_by_sheet_file[sheet_file] = cell_edits

    lookup_wb.close()
    _patch_zip_in_place(output_path, edits_by_sheet_file)

    return output_path, sorted(unplaced_labels)